import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook


LOG_FILE = Path(__file__).resolve().parent.parent / "logs" / "freelance_bot_logs.xlsx"


# ------------------------------------------------------------------
# Jobs sheet — one row per job, reflecting the tiered decision engine.
#
# Compared to the old scoring model, "Score" is gone (there is no
# single number driving the decision anymore) and is replaced with the
# actual evidence trail: which core/supporting keywords fired on each
# side, and the plain-English `reason` the decision table returned.
# This is what makes the log self-explanatory without re-deriving the
# math by hand.
# ------------------------------------------------------------------

JOB_HEADERS = [
    "Timestamp",
    "Job UUID",
    "Job ID",
    "Source",
    "Title",
    "Company",
    "URL",

    "Decision",
    "Decision Reason",

    "Categories",
    "Negative Categories",

    "Has Core Positive",
    "Has Core Negative",
    "Core Positive Hit Count",
    "Supporting Positive Weight",
    "Supporting Negative Weight",

    "Title Core Positive",
    "Title Core Negative",

    "Core Positive Matches",
    "Supporting Positive Matches",
    "Core Negative Matches",
    "Supporting Negative Matches",

    "Hard Reject",
    "Hard Reject Matches",

    "Notify Directly",
    "Needs Gemini",
    "Gemini Decision",
    "Notification Status",
    "Final Decision",
    "Filter Time (ms)",
]


GEMINI_HEADERS = [
    "Timestamp",
    "Job UUID",
    "Decision Before",
    "Reason Before",
    "Prompt Tokens",
    "Completion Tokens",
    "Response Time (ms)",
    "Decision",
    "Confidence",
]


NOTIFICATION_HEADERS = [
    "Timestamp",
    "Job UUID",
    "Platform",
    "Status",
]


ERROR_HEADERS = [
    "Timestamp",
    "Job UUID",
    "Module",
    "Error",
]


COLUMN_MAP = {
    "timestamp": 1,
    "job_uuid": 2,
    "job_id": 3,
    "source": 4,
    "title": 5,
    "company": 6,
    "url": 7,

    "decision": 8,
    "decision_reason": 9,

    "categories": 10,
    "negative_categories": 11,

    "has_core_positive": 12,
    "has_core_negative": 13,
    "core_positive_hit_count": 14,
    "supporting_positive_weight": 15,
    "supporting_negative_weight": 16,

    "title_core_positive": 17,
    "title_core_negative": 18,

    "core_positive_matches": 19,
    "supporting_positive_matches": 20,
    "core_negative_matches": 21,
    "supporting_negative_matches": 22,

    "hard_reject": 23,
    "hard_reject_matches": 24,

    "notify_directly": 25,
    "needs_gemini": 26,
    "gemini_decision": 27,
    "notification_status": 28,
    "final_decision": 29,
    "filter_time_ms": 30,
}


def _join_matches(matches):
    """Render a list of {"keyword", "weight", "category"} dicts as a
    compact, human-readable string for a spreadsheet cell."""
    if not matches:
        return ""
    return ", ".join(
        f"{m['keyword']}({m['weight']}/{m['category']})" for m in matches
    )


class ExcelLogger:

    def __init__(self):
        self.path = LOG_FILE
        self.workbook = None
        self._row_index: dict[str, int] = {}

    def initialize(self):

        self.path.parent.mkdir(parents=True, exist_ok=True)

        if not self.path.exists():

            wb = Workbook()

            wb.remove(wb.active)

            jobs = wb.create_sheet("Jobs")
            jobs.append(JOB_HEADERS)

            gemini = wb.create_sheet("Gemini")
            gemini.append(GEMINI_HEADERS)

            notifications = wb.create_sheet("Notifications")
            notifications.append(NOTIFICATION_HEADERS)

            errors = wb.create_sheet("Errors")
            errors.append(ERROR_HEADERS)

            wb.save(self.path)

        self.workbook = load_workbook(self.path)
        self._build_row_index()

    def _build_row_index(self):

        self._row_index.clear()

        ws = self.workbook["Jobs"]

        for row in range(2, ws.max_row + 1):

            job_uuid = ws.cell(
                row=row,
                column=COLUMN_MAP["job_uuid"],
            ).value

            if job_uuid:
                self._row_index[job_uuid] = row

    def save(self):
        if self.workbook is None:
            return

        temp_path = self.path.with_suffix(".tmp.xlsx")

        self.workbook.save(temp_path)
        os.replace(temp_path, self.path)

    def close(self):
        self.save()
        self.workbook.close()

    def has_job(self, job_uuid) -> bool:
        """Cheap existence check so callers can skip reprocessing a
        job they've already logged (see app.job_processor dedup)."""
        return job_uuid in self._row_index

    def create_job(
        self,
        job_uuid,
        job_id="",
        source="",
        title="",
        company="",
        url="",
        filter_result=None,
        filter_time_ms=None,
        save=True,
    ):
        """
        `filter_result` is expected to be the dict returned by
        `filters.keyword_filter()`. Passing the whole dict (instead of
        a dozen individual keyword arguments) keeps this call in sync
        automatically as the filter's evidence trail evolves.

        `save=False` lets a caller defer the (expensive, full-workbook)
        disk write and batch several updates for the same job into one
        `save()` at the end -- see app.job_processor.process_job.
        """

        filter_result = filter_result or {}

        ws = self.workbook["Jobs"]

        ws.append([
            datetime.now().isoformat(),
            job_uuid,
            job_id,
            source,
            title,
            company,
            url,

            filter_result.get("decision", ""),
            filter_result.get("reason", ""),

            ", ".join(filter_result.get("categories", []) or []),
            ", ".join(filter_result.get("negative_categories", []) or []),

            filter_result.get("has_core_positive", False),
            filter_result.get("has_core_negative", False),
            filter_result.get("core_positive_hit_count", 0),
            filter_result.get("supporting_positive_weight", 0),
            filter_result.get("supporting_negative_weight", 0),

            filter_result.get("title_core_positive", False),
            filter_result.get("title_core_negative", False),

            _join_matches(filter_result.get("positive_core_matches")),
            _join_matches(filter_result.get("positive_supporting_matches")),
            _join_matches(filter_result.get("negative_core_matches")),
            _join_matches(filter_result.get("negative_supporting_matches")),

            filter_result.get("hard_reject", False),
            ", ".join(filter_result.get("hard_reject_matches", []) or []),

            filter_result.get("notify_directly", False),
            filter_result.get("needs_gemini", False),
            "",
            "",
            "",
            filter_time_ms,
        ])

        self._row_index[job_uuid] = ws.max_row

        if save:
            self.save()

    def update_job(self, job_uuid, save=True, **fields):

        row = self._row_index.get(job_uuid)

        if row is None:
            return False

        ws = self.workbook["Jobs"]

        for key, value in fields.items():

            if key not in COLUMN_MAP:
                continue

            if isinstance(value, list):
                value = ", ".join(value)

            ws.cell(
                row=row,
                column=COLUMN_MAP[key],
            ).value = value

        if save:
            self.save()

        return True

    def log_gemini(
        self,
        job_uuid,
        decision_before,
        reason_before,
        prompt_tokens,
        completion_tokens,
        response_time_ms,
        decision,
        confidence,
        save=True,
    ):

        ws = self.workbook["Gemini"]

        ws.append([
            datetime.now().isoformat(),
            job_uuid,
            decision_before,
            reason_before,
            prompt_tokens,
            completion_tokens,
            response_time_ms,
            decision,
            confidence,
        ])

        if save:
            self.save()

    def log_notification(
        self,
        job_uuid,
        platform,
        status,
        save=True,
    ):

        ws = self.workbook["Notifications"]

        ws.append([
            datetime.now().isoformat(),
            job_uuid,
            platform,
            status,
        ])

        if save:
            self.save()

    def log_error(
        self,
        module,
        error,
        job_uuid="",
        save=True,
    ):

        ws = self.workbook["Errors"]

        ws.append([
            datetime.now().isoformat(),
            job_uuid,
            module,
            str(error),
        ])

        if save:
            self.save()


logger = ExcelLogger()


def initialize_workbook():
    logger.initialize()
